#!/usr/bin/env python3
"""
NGX / Fixed Income Dashboard — Auto-Refresh Script
===================================================
Pulls live data from public APIs and websites, updates the Excel workbook,
AND persists a daily snapshot to the database (via db.py) so you build a
time-series history of every fetch.

CHANGED (migration):
  • Persistence now goes through db.py, which targets Postgres in production
    (set DATABASE_URL) and falls back to local SQLite for offline dev. This
    replaces the old hard-coded local ngx.db — that file was being wiped on
    every Streamlit Cloud redeploy, which is why history never accumulated.
  • NTB/FGN rate snapshots are now keyed by their AUCTION date (not "today"),
    so daily runs extend the exact same series that backfill_cbn_history.py
    seeds — no duplicate today-stamped points.

CONFIRMED DATA SOURCES:
  ✅  CBN JSON API        cbn.gov.ng/api/GetAllSecuritiesNTB      (NTB records)
  ✅  CBN JSON API        cbn.gov.ng/api/GetAllSecuritiesFGNBond   (FGN bond records)
  ✅  TradingEconomics    Nigeria 10Y, MPR, CPI, USD 10Y
  ✅  DMO Eurobond Excel  Daily prices + yields for active Eurobonds
  ✅  DMO Auction Page    Latest PDF/Excel auction result link

NOT AVAILABLE (paid subscription required):
  ❌  NGX equity prices   → X-DataPortal  ngxgroup.com/exchange/data
  ❌  FMDQ market data    → info@fmdqgroup.com  +234-1-279-5921

Usage:
    export DATABASE_URL="postgresql+psycopg2://USER:PASS@HOST/DB?sslmode=require"
    python ngx_refresh.py                       # update workbook + DB
    python ngx_refresh.py --backup              # backup workbook before updating
    python ngx_refresh.py --report              # no workbook writes (DB still updates)
    python ngx_refresh.py --workbook /path/to/file.xlsx
    python ngx_refresh.py --no-db               # skip database write
"""

import requests, io, re, json, datetime, argparse, sys, shutil
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import db  # shared Postgres/SQLite layer

# ── CONFIG ────────────────────────────────────────────────────────────────────
WORKBOOK_PATH = "/Users/user/Desktop/NGX_Dashboard/NGX_Stock_Screener_Dashboard.xlsx"
FONT          = "Century Gothic"
AMBER         = "A0740A"
BLACK         = "000000"
MUTED         = "6B7280"
C_INP         = "F0F6FF"
C_SURF        = "F8F8F8"
C_CARD        = "FFFFFF"
TIMEOUT       = 16
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept":     "application/json, text/html, */*",
    "Referer":    "https://www.cbn.gov.ng/",
}

# ── LOGGING ───────────────────────────────────────────────────────────────────
LOG = []
def log(msg, lvl="OK"):
    icons = {"OK":"✓","WARN":"⚠","ERR":"✗","INFO":"·"}
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    line = f"  [{ts}] {icons.get(lvl,'·')}  {msg}"
    print(line); LOG.append(line)

def section(t):
    bar = "═"*62
    print(f"\n{bar}\n  {t}\n{bar}"); LOG.extend([f"\n{bar}", f"  {t}", bar])

# ── CELL HELPERS ──────────────────────────────────────────────────────────────
def _f(sz=10, bold=True, color=BLACK, it=False):
    return Font(name=FONT, size=sz, bold=bold, italic=it, color=color)

def put(ws, r, c, val, nf=None, bg=None, bold=True, color=BLACK, sz=10, ha="center"):
    cell = ws.cell(r, c)
    cell.value = val
    cell.font  = _f(sz, bold, color)
    cell.alignment = Alignment(horizontal=ha, vertical="center")
    if nf: cell.number_format = nf
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def fml(ws, r, c, formula, nf=None, bg=None):
    return put(ws, r, c, formula, nf, bg, bold=False, color=AMBER)

def find_row(ws, substr, col=2, start=1, end=None):
    for r in range(start, (end or ws.max_row)+1):
        v = ws.cell(r, col).value
        if v and substr.lower() in str(v).lower():
            return r
    return None

def get(url):
    return requests.get(url, headers=HEADERS, timeout=TIMEOUT)


# ══════════════════════════════════════════════════════════════════════════════
# PERSISTENCE  —  now via db.py (Postgres in prod, SQLite fallback)
# ══════════════════════════════════════════════════════════════════════════════
#
# Three tables (defined in db.py), all upserted on their primary key so
# re-running the same day overwrites the same row instead of duplicating:
#
#   rates      one row per (snapshot_date, instrument)
#   auctions   one row per (auction_date, security_type, tenor)
#   eurobonds  one row per (snapshot_date, bond_name)
#
def _iso(s):
    """Convert 'dd/mm/yyyy' string to 'YYYY-MM-DD'. Returns None on failure."""
    try:
        return datetime.datetime.strptime(s, "%d/%m/%Y").date().isoformat()
    except Exception:
        return None


def _rate_rows(ntb, fgn_bonds, bench, now):
    """Build rows for the rates table.

    NTB/FGN points are stamped with their AUCTION date so they extend the same
    series produced by the history backfill. Benchmarks (MPR/CPI/10Y/USD10Y)
    have no auction date, so they're stamped with today's date.
    """
    today = datetime.date.today().isoformat()
    rows = []
    for tenor, rec in ntb.items():
        d = _iso(rec.get("auction_date", "")) or today
        rows.append(dict(snapshot_date=d, instrument=f"NTB_{tenor}",
                         rate=float(rec["rate"]), source="CBN", fetched_at=now))
    for tenor, rec in fgn_bonds.items():
        d = _iso(rec.get("auction_date", "")) or today
        rows.append(dict(snapshot_date=d, instrument=f"FGN_{tenor}",
                         rate=float(rec["rate"]), source="CBN", fetched_at=now))
    for key, val in bench.items():
        if val is not None:
            rows.append(dict(snapshot_date=today, instrument=key,
                             rate=float(val), source="TE", fetched_at=now))
    return rows


def _auction_rows(ntb, fgn_bonds, now):
    rows = []
    for tenor, rec in ntb.items():
        ad = _iso(rec.get("auction_date", ""))
        if ad:
            rows.append(dict(auction_date=ad, security_type="NTB", tenor=tenor,
                             offered=rec.get("offered"), subscribed=rec.get("subscribed"),
                             allotted=rec.get("allotted"), stop_rate=rec.get("rate"),
                             maturity_date=_iso(rec.get("maturity", "")), fetched_at=now))
    for tenor, rec in fgn_bonds.items():
        ad = _iso(rec.get("auction_date", ""))
        if ad:
            rows.append(dict(auction_date=ad, security_type="FGN", tenor=tenor,
                             offered=rec.get("offered"), subscribed=rec.get("subscribed"),
                             allotted=rec.get("allotted"), stop_rate=rec.get("rate"),
                             maturity_date=_iso(rec.get("maturity", "")), fetched_at=now))
    return rows


def _eurobond_rows(eu, now):
    if not eu or not eu.get("date") or not eu.get("prices"):
        return []
    snap = (eu["date"].isoformat() if hasattr(eu["date"], "isoformat") else str(eu["date"]))
    rows = []
    for name, price in eu["prices"].items():
        yld = eu["yields"].get(name)
        rows.append(dict(snapshot_date=snap, bond_name=name,
                         price=float(price) if price is not None else None,
                         yield_pct=float(yld) if yld is not None else None,
                         fetched_at=now))
    return rows


def persist_to_db(ntb, fgn_bonds, bench, eu):
    """Initialise schema then upsert everything. Wrapped so a DB failure never
    blocks the Excel update."""
    section("PERSISTING TO DATABASE")
    log(f"Target: {db.database_url().split('@')[-1]}")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    try:
        db.init_schema()
        eng = db.get_engine()
        n_rates = db.upsert(eng, "rates",     _rate_rows(ntb, fgn_bonds, bench, now))
        n_auc   = db.upsert(eng, "auctions",  _auction_rows(ntb, fgn_bonds, now))
        n_eu    = db.upsert(eng, "eurobonds", _eurobond_rows(eu, now))
        log(f"  rates:     {n_rates} rows upserted")
        log(f"  auctions:  {n_auc} rows upserted")
        log(f"  eurobonds: {n_eu} rows upserted")
        for tbl in ("rates", "auctions", "eurobonds"):
            cnt = db.read_sql(f"SELECT COUNT(*) AS n FROM {tbl}")
            total = int(cnt.iloc[0]["n"]) if not cnt.empty else 0
            log(f"  total in {tbl}: {total} rows", "INFO")
        return n_rates + n_auc + n_eu
    except Exception as e:
        log(f"DB write failed (Excel will still update): {e}", "ERR")
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 1 — CBN NTB AUCTION API
# ══════════════════════════════════════════════════════════════════════════════
def fetch_cbn_ntb():
    section("SOURCE 1 — CBN NTB Auction API")
    try:
        r = get("https://www.cbn.gov.ng/api/GetAllSecuritiesNTB")
        data = r.json()
        log(f"CBN NTB API: {len(data)} total records")

        def pdate(s):
            try: return datetime.datetime.strptime(s, "%d/%m/%Y")
            except: return datetime.datetime(1900,1,1)

        data.sort(key=lambda x: pdate(x.get("auctionDate","")), reverse=True)

        result = {}
        for rec in data:
            tenor = rec.get("tenor","").upper().replace(" ","").replace("-","")
            key = ("91D" if "91" in tenor else
                   "182D" if ("182" in tenor or "180" in tenor) else
                   "364D" if ("364" in tenor or "365" in tenor) else None)
            if key and key not in result:
                rate_str = rec.get("rate","0") or "0"
                try:
                    rate = float(rate_str)
                    if rate > 0:
                        result[key] = {
                            "rate":        round(rate/100, 6),
                            "auction_date":rec.get("auctionDate",""),
                            "maturity":    rec.get("maturityDate",""),
                            "offered":     float(rec.get("amtOffered",0) or 0),
                            "subscribed":  float(rec.get("totalSubscription",0) or 0),
                            "allotted":    float(rec.get("totalSuccessful",0) or 0),
                        }
                except:
                    pass
            if len(result) == 3:
                break

        for k,v in result.items():
            log(f"  {k}: {v['rate']*100:.2f}%  (auction {v['auction_date']})")
        return result
    except Exception as e:
        log(f"CBN NTB API failed: {e}", "ERR")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 2 — CBN FGN BOND AUCTION API
# ══════════════════════════════════════════════════════════════════════════════
def fetch_cbn_fgn():
    section("SOURCE 2 — CBN FGN Bond Auction API")
    try:
        r = get("https://www.cbn.gov.ng/api/GetAllSecuritiesFGNBond")
        data = r.json()
        log(f"CBN FGN API: {len(data)} total records")

        def pdate(s):
            try: return datetime.datetime.strptime(s, "%d/%m/%Y")
            except: return datetime.datetime(1900,1,1)

        data.sort(key=lambda x: pdate(x.get("auctionDate","")), reverse=True)

        TENOR_MAP = {
            "2YEAR":"2Y","2YR":"2Y","3YEAR":"3Y","5YEAR":"5Y","5YR":"5Y",
            "7YEAR":"7Y","7YR":"7Y","10YEAR":"10Y","10YR":"10Y",
            "15YEAR":"15Y","20YEAR":"20Y","30YEAR":"30Y",
        }
        result = {}
        for rec in data:
            raw = rec.get("tenor","").upper().replace(" ","").replace("-","")
            key = TENOR_MAP.get(raw)
            if key and key not in result:
                rate_str = rec.get("rate","0") or "0"
                try:
                    rate = float(rate_str)
                    if rate > 0:
                        result[key] = {
                            "rate":        round(rate/100, 6),
                            "auction_date":rec.get("auctionDate",""),
                            "maturity":    rec.get("maturityDate",""),
                            "offered":     float(rec.get("amtOffered",0) or 0),
                            "subscribed":  float(rec.get("totalSubscription",0) or 0),
                            "allotted":    float(rec.get("totalSuccessful",0) or 0),
                        }
                except:
                    pass

        for k,v in result.items():
            log(f"  {k}: {v['rate']*100:.2f}%  (auction {v['auction_date']})")
        return result
    except Exception as e:
        log(f"CBN FGN API failed: {e}", "ERR")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 3 — TRADINGECONOMICS (10Y, MPR, CPI, USD10Y)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_te(url):
    """Two-stage TE extractor: JSON 'last' field, then meta description."""
    try:
        r = get(url)
        m = re.search(r'"last"\s*:\s*([\d.]+)', r.text[:20000])
        if m: return float(m.group(1))
        m2 = re.search(r'content="[^"]*?(\d{1,3}(?:\.\d+)?)\s*percent', r.text[:3000], re.IGNORECASE)
        if m2: return float(m2.group(1))
        m3 = re.search(r'(?:at|steady at|rate is)\s*([\d]+\.[\d]+)%', r.text[:3000], re.IGNORECASE)
        if m3: return float(m3.group(1))
    except: pass
    return None


def fetch_benchmarks():
    section("SOURCE 3 — TradingEconomics Benchmark Rates")
    result = {}
    checks = [
        ("Nigeria 10Y",  "https://tradingeconomics.com/nigeria/government-bond-yield", "10Y"),
        ("CBN MPR",      "https://tradingeconomics.com/nigeria/interest-rate",          "MPR"),
        ("Nigeria CPI",  "https://tradingeconomics.com/nigeria/inflation-cpi",         "CPI"),
        ("US 10Y UST",   "https://tradingeconomics.com/united-states/government-bond-yield","USD10Y"),
    ]
    for label, url, key in checks:
        val = fetch_te(url)
        if val:
            result[key] = round(val/100, 6) if val > 1 else round(val, 6)
            log(f"  {label}: {val:.3f}%")
        else:
            log(f"  {label}: FAILED", "WARN")
    return result


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 4 — DMO EUROBOND EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def find_dmo_eurobond_url():
    """Scrape DMO page to find latest cumulative Eurobond Excel download."""
    try:
        r = get("https://www.dmo.gov.ng/fgn-bonds/eurobonds-trading")
        soup = BeautifulSoup(r.text, "lxml")
        best = ("", 0)
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "eurobond" in href.lower() and "/file" in href.lower():
                if "daily" in href.lower() or "from" in href.lower():
                    m = re.search(r'/(\d+)-', href)
                    if m and int(m.group(1)) > best[1]:
                        best = (href, int(m.group(1)))
        if not best[0]:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "eurobond" in href.lower() and "/file" in href.lower():
                    m = re.search(r'/(\d+)-', href)
                    if m and int(m.group(1)) > best[1]:
                        best = (href, int(m.group(1)))
        if best[0]:
            return f"https://www.dmo.gov.ng{best[0]}" if best[0].startswith("/") else best[0]
    except Exception as e:
        log(f"DMO page scrape: {e}", "WARN")
    return None


def fetch_eurobonds():
    section("SOURCE 4 — DMO Eurobond Excel")
    url = find_dmo_eurobond_url()
    if not url:
        log("No Eurobond file URL found", "ERR")
        return None
    log(f"URL: {url[-65:]}")
    try:
        r = get(url)
        ct = r.headers.get("content-type","")
        if "spreadsheet" not in ct and "excel" not in ct and "openxmlformats" not in ct:
            log(f"Unexpected content-type: {ct}", "WARN")
            if len(r.content) < 5000:
                return None

        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        out = {"date": None, "prices": {}, "yields": {}}

        for sheet_name, store_key in [("PRICE","prices"),("YIELD","yields")]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            hdrs = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            last = None
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and isinstance(row[0], datetime.datetime):
                    last = row
            if not last: continue
            if store_key == "prices" and last[0]:
                out["date"] = last[0].date() if hasattr(last[0],"date") else last[0]
            for i, h in enumerate(hdrs[1:], 1):
                if not h or "US$" not in str(h): continue
                v = last[i] if i < len(last) else None
                if not isinstance(v, (int, float)) or v <= 0: continue
                name = str(h).replace("\n"," ").strip()[:50]
                if store_key == "prices":
                    out["prices"][name] = round(float(v), 4)
                else:
                    if v > 50:
                        out["yields"][name] = round(v/10000, 6)
                    elif v > 1:
                        out["yields"][name] = round(v/100, 6)
                    else:
                        out["yields"][name] = round(float(v), 6)

        log(f"Eurobonds: {len(out['prices'])} price records, "
            f"{len(out['yields'])} yield records, date={out['date']}")
        return out if out["date"] else None
    except Exception as e:
        log(f"Eurobond parse error: {e}", "ERR")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE 5 — DMO LATEST AUCTION LINK
# ══════════════════════════════════════════════════════════════════════════════
def fetch_latest_auction_url():
    section("SOURCE 5 — DMO Latest Bond Auction Link")
    try:
        r = get("https://www.dmo.gov.ng/fgn-bonds/bonds-auction-results")
        soup = BeautifulSoup(r.text, "lxml")
        best = ("", 0)
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "bonds-auction-results" in href and "/file" in href:
                m = re.search(r'/(\d+)-', href)
                if m and int(m.group(1)) > best[1]:
                    best = (href, int(m.group(1)))
        if best[0]:
            url = f"https://www.dmo.gov.ng{best[0]}" if best[0].startswith("/") else best[0]
            log(f"Latest auction: ...{best[0][-55:]}")
            return url
        log("No auction link found", "WARN")
    except Exception as e:
        log(f"Auction link: {e}", "ERR")
    return None


# ══════════════════════════════════════════════════════════════════════════════
# WRITE — CBNRates Sheet
# ══════════════════════════════════════════════════════════════════════════════
def write_cbn_rates(wb, ntb, fgn_bonds, bench):
    section("WRITING: CBNRates Sheet")
    ws = wb["CBNRates"]

    updates = {}
    if bench.get("MPR"):    updates["MPR"]     = bench["MPR"]
    if ntb.get("91D"):      updates["91-Day"]  = ntb["91D"]["rate"]
    if ntb.get("91D"):      updates["91  Day"] = ntb["91D"]["rate"]
    if ntb.get("182D"):     updates["182-Day"] = ntb["182D"]["rate"]
    if ntb.get("182D"):     updates["182  Day"]= ntb["182D"]["rate"]
    if ntb.get("364D"):     updates["364-Day"] = ntb["364D"]["rate"]
    if ntb.get("364D"):     updates["364  Day"]= ntb["364D"]["rate"]
    if fgn_bonds.get("2Y"):  updates["2-Year"]  = fgn_bonds["2Y"]["rate"]
    if fgn_bonds.get("2Y"):  updates["2  Year"] = fgn_bonds["2Y"]["rate"]
    if fgn_bonds.get("5Y"):  updates["5-Year"]  = fgn_bonds["5Y"]["rate"]
    if fgn_bonds.get("5Y"):  updates["5  Year"] = fgn_bonds["5Y"]["rate"]
    if bench.get("10Y") or fgn_bonds.get("10Y"):
        v = bench.get("10Y") or fgn_bonds["10Y"]["rate"]
        updates["10-Year"] = v; updates["10  Year"] = v
    if fgn_bonds.get("15Y"): updates["15-Year"] = fgn_bonds["15Y"]["rate"]
    if fgn_bonds.get("15Y"): updates["15  Year"]= fgn_bonds["15Y"]["rate"]
    if fgn_bonds.get("20Y"): updates["20-Year"] = fgn_bonds["20Y"]["rate"]
    if fgn_bonds.get("20Y"): updates["20  Year"]= fgn_bonds["20Y"]["rate"]
    if fgn_bonds.get("30Y"): updates["30-Year"] = fgn_bonds["30Y"]["rate"]
    if fgn_bonds.get("30Y"): updates["30  Year"]= fgn_bonds["30Y"]["rate"]
    if bench.get("CPI"):    updates["Inflation"] = bench["CPI"]
    if bench.get("CPI"):    updates["CPI"]       = bench["CPI"]

    written = set()
    count = 0
    for label, new_val in updates.items():
        if label in written: continue
        row = find_row(ws, label)
        if not row: continue
        curr = ws.cell(row, 3).value
        if isinstance(curr, (int, float)) and curr > 0:
            put(ws, row, 4, round(float(curr),6), "0.00%",
                C_SURF if row%2 else C_CARD, True, BLACK)
        cell = ws.cell(row, 3)
        cell.value = round(new_val, 6)
        cell.number_format = "0.00%"
        cell.font = _f(10, True, BLACK)
        cell.fill = PatternFill("solid", fgColor=C_INP)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        log(f"  [{row}] {label.strip()}: → {new_val*100:.3f}%")
        written.add(label); count += 1

    note = find_row(ws, "auto-refresh") or find_row(ws, "last updated") or ws.max_row+2
    ws.cell(note, 2).value = (
        f"  Auto-refreshed: {datetime.datetime.now().strftime('%d %b %Y  %H:%M')}  ·  "
        f"Sources: CBN API (NTB/FGN) · TradingEconomics (10Y/MPR/CPI) · DMO Eurobond Excel")
    ws.cell(note, 2).font = _f(8, False, MUTED, True)
    log(f"CBNRates: {count} cells updated")
    return count


# ══════════════════════════════════════════════════════════════════════════════
# WRITE — Yield Curve NTB Auction Log
# ══════════════════════════════════════════════════════════════════════════════
def write_yield_curve(wb, ntb):
    section("WRITING: Yield Curve — NTB Auction Log")
    ws = wb["Yield Curve"]
    section_row = find_row(ws, "NTB  AUCTION TRACKER") or find_row(ws, "NTB AUCTION")
    if not section_row:
        log("NTB auction section not found", "WARN")
        return 0

    data_start = section_row + 2
    count = 0

    for tenor_key, tenor_label in [("91D","91 Day"),("182D","182 Day"),("364D","364 Day")]:
        rec = ntb.get(tenor_key)
        if not rec: continue
        try:
            dt = datetime.datetime.strptime(rec["auction_date"], "%d/%m/%Y").date()
            mat_dt = datetime.datetime.strptime(rec["maturity"], "%d/%m/%Y").date()
        except:
            continue

        already = False
        for r in range(data_start, min(data_start+50, ws.max_row+1)):
            bv = ws.cell(r,2).value
            if bv and str(bv).strip() == str(dt):
                tv = ws.cell(r,3).value
                if tv and tenor_label.lower() in str(tv).lower():
                    already = True; break

        if already:
            log(f"  {tenor_key} {rec['auction_date']}: already logged")
            continue

        ws.insert_rows(data_start)
        bg = C_CARD
        put(ws, data_start, 2, dt,                     "YYYY-MM-DD", bg)
        put(ws, data_start, 3, tenor_label,            "@",          bg, ha="center")
        put(ws, data_start, 4, rec["offered"]/1000,    "#,##0.0",    bg)
        put(ws, data_start, 5, rec["subscribed"]/1000, "#,##0.0",    bg)
        put(ws, data_start, 6, rec["allotted"]/1000,   "#,##0.0",    bg)
        put(ws, data_start, 7, rec["rate"],             "0.00%",      C_INP)
        put(ws, data_start, 8, mat_dt,                 "YYYY-MM-DD", bg)
        ws.row_dimensions[data_start].height = 15
        log(f"  Added {tenor_key}: {rec['auction_date']} → {rec['rate']*100:.2f}%")
        count += 1

    log(f"Yield Curve NTB log: {count} new rows added")
    return count


# ══════════════════════════════════════════════════════════════════════════════
# WRITE — Eurobonds Sheet
# ══════════════════════════════════════════════════════════════════════════════
def write_eurobonds(wb, eu):
    section("WRITING: Eurobonds Sheet")
    if not eu or not eu.get("prices"):
        log("No Eurobond data to write", "WARN"); return 0

    ws = wb.get("Eurobonds") if hasattr(wb, "get") else (wb["Eurobonds"] if "Eurobonds" in wb.sheetnames else None)
    if not ws:
        log("Eurobonds sheet missing", "ERR"); return 0

    count = 0
    lookup = {}
    for dmo_name, price in eu["prices"].items():
        m = re.search(r'^([\d.]+)%', dmo_name)
        if m:
            coupon = float(m.group(1))
            lookup[coupon] = {"price": price, "yield": eu["yields"].get(dmo_name, 0), "name": dmo_name}

    for r in range(7, ws.max_row+1):
        coupon_raw = ws.cell(r, 3).value
        if not coupon_raw: continue
        try:
            coupon_val = float(str(coupon_raw).strip().rstrip("%"))
            coupon_pct = coupon_val * 100 if coupon_val < 1 else coupon_val
        except: continue

        for k, v in lookup.items():
            if abs(k - coupon_pct) < 0.02:
                price = v["price"]; yld = v["yield"]
                if price > 0:
                    put(ws, r, 8, round(price,3), "0.000", C_INP)
                if yld > 0:
                    put(ws, r, 10, round(yld,6), "0.00%", C_INP)
                count += 1; break

    for r in range(2,6):
        v = ws.cell(r,2).value
        if v and "Source:" in str(v):
            ws.cell(r,2).value = (
                f"Source: DMO Eurobond daily Excel  ·  "
                f"Data as at: {eu['date'].strftime('%d %b %Y') if eu.get('date') else 'N/A'}  ·  "
                f"Auto-refreshed: {datetime.datetime.now().strftime('%d %b %Y  %H:%M')}"
            )
            ws.cell(r,2).font = _f(8, False, MUTED, True)
            break

    log(f"Eurobonds: {count} bonds updated (data date: {eu.get('date')})")
    return count


# ══════════════════════════════════════════════════════════════════════════════
# WRITE — FI Dashboard refresh note
# ══════════════════════════════════════════════════════════════════════════════
def write_fi_dashboard_note(wb, eu, bench, ntb):
    ws = wb["FI Dashboard"]
    note_row = (find_row(ws, "Auto-refreshed") or
                find_row(ws, "links live from") or
                find_row(ws, "Data in this dashboard"))
    if note_row:
        ws.cell(note_row, 2).value = (
            f"  Auto-refreshed: {datetime.datetime.now().strftime('%d %b %Y  %H:%M')}  ·  "
            f"NTBs: CBN API (latest: {ntb.get('91D',{}).get('auction_date','N/A')})  ·  "
            f"10Y FGN: {(bench.get('10Y',0)*100):.3f}%  ·  "
            f"Eurobonds: DMO Excel (to {eu.get('date','N/A') if eu else 'N/A'})  ·  "
            f"NGX prices: manual (X-DataPortal required)"
        )
        ws.cell(note_row, 2).font = _f(8, False, MUTED, True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="NGX/FI Dashboard Refresh")
    ap.add_argument("--backup",   action="store_true")
    ap.add_argument("--report",   action="store_true", help="Skip workbook update (DB still persists)")
    ap.add_argument("--workbook", default=WORKBOOK_PATH)
    ap.add_argument("--no-db",    action="store_true", help="Skip database write")
    args = ap.parse_args()

    print(f"\n{'═'*62}")
    print(f"  NGX / Fixed Income Dashboard — Auto Refresh")
    print(f"  {datetime.datetime.now().strftime('%A, %d %B %Y  %H:%M:%S')}")
    print(f"{'═'*62}")

    # FETCH
    ntb        = fetch_cbn_ntb()
    fgn_bonds  = fetch_cbn_fgn()
    bench      = fetch_benchmarks()
    eu         = fetch_eurobonds()
    auction    = fetch_latest_auction_url()

    # SUMMARY
    section("DATA FETCH SUMMARY")
    items = [
        ("NTB 91D stop rate",    ntb.get("91D"),        f"{ntb['91D']['rate']*100:.2f}%  (auction {ntb['91D']['auction_date']})" if ntb.get("91D") else None),
        ("NTB 182D stop rate",   ntb.get("182D"),       f"{ntb['182D']['rate']*100:.2f}%  (auction {ntb['182D']['auction_date']})" if ntb.get("182D") else None),
        ("NTB 364D stop rate",   ntb.get("364D"),       f"{ntb['364D']['rate']*100:.2f}%  (auction {ntb['364D']['auction_date']})" if ntb.get("364D") else None),
        ("FGN Bond 5Y",          fgn_bonds.get("5Y"),   f"{fgn_bonds['5Y']['rate']*100:.2f}%  (auction {fgn_bonds['5Y']['auction_date']})" if fgn_bonds.get("5Y") else None),
        ("FGN Bond 10Y",         fgn_bonds.get("10Y") or bench.get("10Y"), f"{(bench.get('10Y',0))*100:.3f}%" if bench.get("10Y") else None),
        ("CBN MPR",              bench.get("MPR"),      f"{bench['MPR']*100:.2f}%" if bench.get("MPR") else None),
        ("Nigeria CPI",          bench.get("CPI"),      f"{bench['CPI']*100:.1f}%" if bench.get("CPI") else None),
        ("US 10Y Treasury",      bench.get("USD10Y"),   f"{bench['USD10Y']*100:.3f}%" if bench.get("USD10Y") else None),
        ("Eurobonds (DMO Excel)", eu is not None,       f"{len(eu['prices'])} bonds, data to {eu['date']}" if eu else None),
        ("Latest DMO auction",   auction is not None,   (auction or "")[-55:] if auction else None),
        ("NGX equity prices",    False,                 "X-DataPortal subscription required"),
        ("FMDQ CP/Bond data",    False,                 "FMDQ data services subscription required"),
    ]
    for label, ok, detail in items:
        icon = "✅" if ok else "❌"
        det  = f"  →  {detail}" if detail else ""
        print(f"    {icon}  {label:28}{det}")

    # PERSIST (before workbook write — so DB updates even in --report mode)
    if not args.no_db:
        persist_to_db(ntb, fgn_bonds, bench, eu)

    if args.report:
        print("\n  Report-only mode — no workbook changes written.")
        return

    # LOAD & UPDATE WORKBOOK
    section("UPDATING WORKBOOK")
    if not Path(args.workbook).exists():
        log(f"File not found: {args.workbook}", "ERR"); sys.exit(1)

    if args.backup:
        bk = args.workbook.replace(".xlsx",
            f"_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        shutil.copy2(args.workbook, bk)
        log(f"Backup: {Path(bk).name}")

    wb = openpyxl.load_workbook(args.workbook)
    log(f"Loaded: {Path(args.workbook).name}  ({len(wb.sheetnames)} sheets)")

    total = 0
    total += write_cbn_rates(wb, ntb, fgn_bonds, bench)
    total += write_yield_curve(wb, ntb)
    if eu:
        total += write_eurobonds(wb, eu)
    write_fi_dashboard_note(wb, eu, bench, ntb)

    wb.save(args.workbook)
    log(f"Saved: {Path(args.workbook).name}")

    section("DONE")
    print(f"""
  Total cells updated: {total}
  Latest NTB auction:  {ntb.get('91D',{}).get('auction_date','N/A')}
  Latest DMO PDF:      {auction or 'Not found'}
  Eurobond data date:  {eu.get('date') if eu else 'N/A'}
  Database:            {db.database_url().split('@')[-1]}
""")

if __name__ == "__main__":
    main()
